"""
CyberGuard - Δήμος Αθηναίων Database Integration Module
Σύνδεση με τη βάση δεδομένων του Δήμου Αθηναίων
"""

import json
import csv
import os
from datetime import datetime
from typing import List, Dict, Optional

class MunicipalDatabase:
    """Interface for Δήμος Αθηναίων database"""
    
    def __init__(self, db_path: str = None):
        """
        Initialize database connection
        
        Args:
            db_path: Path to database file (CSV, JSON, or Excel)
        """
        self.db_path = db_path or "C:\\Users\\User\\Desktop\\dimos_athens_db.json"
        self.employees = []
        self.departments = []
        self.safe_email_domains = []
        self.verified_senders = []
        
        if db_path:
            self.load_database()
    
    def load_database(self):
        """Load database from file"""
        if not os.path.exists(self.db_path):
            self.create_default_database()
            return
        
        try:
            if self.db_path.endswith('.json'):
                self.load_json_database()
            elif self.db_path.endswith('.csv'):
                self.load_csv_database()
            elif self.db_path.endswith(('.xlsx', '.xls')):
                self.load_excel_database()
        except Exception as e:
            print(f"Error loading database: {e}")
            self.create_default_database()
    
    def load_json_database(self):
        """Load from JSON file"""
        try:
            with open(self.db_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                self.employees = data.get('employees', [])
                self.departments = data.get('departments', [])
                self.safe_email_domains = data.get('safe_email_domains', [])
                self.verified_senders = data.get('verified_senders', [])
            print(f"✅ Loaded {len(self.employees)} employees from database")
        except Exception as e:
            print(f"❌ Error loading JSON: {e}")
    
    def load_csv_database(self):
        """Load from CSV file"""
        try:
            with open(self.db_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                self.employees = list(reader)
            print(f"✅ Loaded {len(self.employees)} employees from CSV")
        except Exception as e:
            print(f"❌ Error loading CSV: {e}")
    
    def load_excel_database(self):
        """Load from Excel file"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.db_path)
            ws = wb.active
            
            # Get header
            headers = [cell.value for cell in ws[1]]
            
            # Get data
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Skip empty rows
                    self.employees.append(dict(zip(headers, row)))
            
            print(f"✅ Loaded {len(self.employees)} employees from Excel")
        except ImportError:
            print("⚠️  openpyxl not installed. Use CSV instead.")
        except Exception as e:
            print(f"❌ Error loading Excel: {e}")
    
    def create_default_database(self):
        """Create default database with sample data"""
        default_data = {
            'employees': [
                {
                    'id': 'DAA001',
                    'name': 'Γιάννης Παπαδόπουλος',
                    'email': 'gpapadopoulos@dimos-athens.gr',
                    'department': 'Διοίκηση',
                    'position': 'Δήμαρχος',
                    'phone': '210-1234567'
                },
                {
                    'id': 'DAA002',
                    'name': 'Μαρία Σταματοπούλου',
                    'email': 'mstamatopoulou@dimos-athens.gr',
                    'department': 'IT',
                    'position': 'Διευθύντρια IT',
                    'phone': '210-1234568'
                },
                {
                    'id': 'DAA003',
                    'name': 'Κώστας Νικολάου',
                    'email': 'knikolaou@dimos-athens.gr',
                    'department': 'Αδειοδοτήσεις',
                    'position': 'Υπάλληλος',
                    'phone': '210-1234569'
                }
            ],
            'departments': [
                {'name': 'Διοίκηση', 'code': 'ADM', 'email': 'admin@dimos-athens.gr'},
                {'name': 'IT', 'code': 'IT', 'email': 'it@dimos-athens.gr'},
                {'name': 'Αδειοδοτήσεις', 'code': 'LIC', 'email': 'licenses@dimos-athens.gr'},
                {'name': 'Φιλοδοξία', 'code': 'TAX', 'email': 'taxes@dimos-athens.gr'},
                {'name': 'Υπηρεσίες Πολίτη', 'code': 'CIT', 'email': 'citizen@dimos-athens.gr'}
            ],
            'safe_email_domains': [
                'dimos-athens.gr',
                'mail.dimos-athens.gr',
                'athens.gr',
                'gov.gr'
            ],
            'verified_senders': [
                'admin@dimos-athens.gr',
                'it@dimos-athens.gr',
                'licenses@dimos-athens.gr'
            ]
        }
        
        # Save default database
        try:
            with open(self.db_path, 'w', encoding='utf-8') as f:
                json.dump(default_data, f, ensure_ascii=False, indent=2)
            
            self.employees = default_data['employees']
            self.departments = default_data['departments']
            self.safe_email_domains = default_data['safe_email_domains']
            self.verified_senders = default_data['verified_senders']
            
            print(f"✅ Created default database with {len(self.employees)} employees")
        except Exception as e:
            print(f"❌ Error creating database: {e}")
    
    def verify_employee_email(self, email: str) -> Optional[Dict]:
        """
        Verify if email belongs to a known employee
        
        Args:
            email: Email address to verify
            
        Returns:
            Employee data if found, None otherwise
        """
        for emp in self.employees:
            if emp.get('email', '').lower() == email.lower():
                return emp
        return None
    
    def verify_email_domain(self, email: str) -> bool:
        """
        Verify if email domain is from Δήμος Αθηναίων
        
        Args:
            email: Email address to check
            
        Returns:
            True if domain is in safe list
        """
        domain = email.split('@')[-1].lower()
        # Ensure comparison is case-insensitive
        return domain in [d.lower() for d in self.safe_email_domains]
    
    def get_department_info(self, dept_name: str) -> Optional[Dict]:
        """Get department information"""
        for dept in self.departments:
            if dept.get('name', '').lower() == dept_name.lower():
                return dept
        return None
    
    def get_department_emails(self, dept_name: str) -> List[str]:
        """Get all emails from a department"""
        emails = []
        for emp in self.employees:
            if emp.get('department', '').lower() == dept_name.lower():
                emails.append(emp.get('email'))
        return emails
    
    def is_verified_sender(self, email: str) -> bool:
        """Check if sender is verified"""
        return email.lower() in [v.lower() for v in self.verified_senders]
    
    def add_employee(self, employee_data: Dict) -> bool:
        """Add new employee to database"""
        try:
            self.employees.append(employee_data)
            self.save_database()
            return True
        except Exception as e:
            print(f"Error adding employee: {e}")
            return False
    
    def update_employee(self, employee_id: str, updates: Dict) -> bool:
        """Update employee information"""
        try:
            for emp in self.employees:
                if emp.get('id') == employee_id:
                    emp.update(updates)
                    self.save_database()
                    return True
            return False
        except Exception as e:
            print(f"Error updating employee: {e}")
            return False
    
    def save_database(self) -> bool:
        """Save database to file"""
        try:
            data = {
                'employees': self.employees,
                'departments': self.departments,
                'safe_email_domains': self.safe_email_domains,
                'verified_senders': self.verified_senders
            }
            
            with open(self.db_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            return True
        except Exception as e:
            print(f"Error saving database: {e}")
            return False
    
    def get_statistics(self) -> Dict:
        """Get database statistics"""
        return {
            'total_employees': len(self.employees),
            'total_departments': len(self.departments),
            'safe_domains': len(self.safe_email_domains),
            'verified_senders': len(self.verified_senders),
            'last_updated': datetime.now().isoformat()
        }


# ==================== Enhanced Detection with Municipal Data ====================

def municipal_based_detection(email_text: str, email_from: str, database: MunicipalDatabase) -> tuple:
    """
    Phishing detection using municipal database
    
    Args:
        email_text: Email content
        email_from: Sender email address
        database: MunicipalDatabase instance
        
    Returns:
        Tuple of (score, issues)
    """
    score = 0.0
    issues = []
    
    # Check if sender is from municipal domain
    if database.verify_email_domain(email_from):
        score -= 0.30  # Reduce risk if from municipal domain
        employee = database.verify_employee_email(email_from)
        if employee:
            issues.append(f"✅ Verified: {employee.get('name', 'Unknown')} ({employee.get('position')})")
    else:
        score += 0.15  # Increase risk if NOT from municipal domain
        issues.append(f"⚠️  Unknown domain: {email_from.split('@')[-1]}")
    
    # Check if sender is verified
    if database.is_verified_sender(email_from):
        score -= 0.20
        issues.append("✅ Verified sender")
    
    # Check for impersonation attempts (common names spoofed)
    for emp in database.employees:
        if emp.get('name', '').lower() in email_text.lower():
            # Check if domain is correct
            if not email_from.endswith('@dimos-athens.gr'):
                score += 0.35
                issues.append(f"🚨 Possible impersonation of {emp.get('name')}")
    
    return min(score, 1.0), issues


# ==================== Test ====================

if __name__ == '__main__':
    # Test database
    db = MunicipalDatabase()
    
    print("\n" + "="*50)
    print("Δήμος Αθηναίων Database - Test")
    print("="*50)
    
    # Statistics
    stats = db.get_statistics()
    print(f"\n📊 Database Statistics:")
    for key, value in stats.items():
        print(f"  {key}: {value}")
    
    # Test email verification
    test_emails = [
        'gpapadopoulos@dimos-athens.gr',
        'someone@gmail.com',
        'unknown@dimos-athens.gr'
    ]
    
    print(f"\n✉️  Email Verification:")
    for email in test_emails:
        emp = db.verify_employee_email(email)
        if emp:
            print(f"  ✅ {email} -> {emp.get('name')}")
        else:
            print(f"  ❌ {email} -> Not found")
    
    # Test domain verification
    print(f"\n🌐 Domain Verification:")
    for email in test_emails:
        domain_ok = db.verify_email_domain(email)
        print(f"  {email}: {'✅ Safe' if domain_ok else '❌ Unsafe'}")
    
    # Test department lookup
    print(f"\n🏛️  Department Info:")
    dept = db.get_department_info('IT')
    if dept:
        print(f"  {dept.get('name')}: {dept.get('email')}")
        dept_emails = db.get_department_emails('IT')
        for email in dept_emails:
            print(f"    - {email}")
